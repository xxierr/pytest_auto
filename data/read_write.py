import openpyxl
from config.config import datafile_p


class ReadWrite:
    caseid = "用例编号"
    casemodule = "模块"
    casedescription = "用例说明"
    requestaddress = "请求地址"
    requestmethod = "请求方式"
    requestparam = "请求参数"
    requestheader = "请求头"
    precondition = "前置条件"
    isexec = "是否执行"
    status = "状态码"
    expectresult = "期望结果"
    actualresult = "实际结果"
    title = [caseid, casemodule, casedescription, requestaddress, requestmethod, requestparam, requestheader,
             precondition, isexec, status, expectresult, actualresult]

    def __init__(self, sheet_num):
        self.sheet_num = sheet_num
        self.title = ReadWrite.title

    def read(self):
        wb = openpyxl.load_workbook(datafile_p)
        sheets = wb.sheetnames
        sheetname = wb[sheets[self.sheet_num - 1]]
        wb.active = sheetname
        nrows = sheetname.max_row
        ncols = sheetname.max_column
        col_value = []
        case_values = []
        for i in range(2, nrows + 1):
            for j in range(1, ncols + 1):
                col_value.append(sheetname.cell(i, j).value)
            row_values = dict(zip(self.title, col_value))
            col_value = []
            case_values.append(row_values)
        return case_values

    def write(self, caseid, actual):
        wb = openpyxl.load_workbook(datafile_p)
        sheets = wb.sheetnames
        sheetname = wb[sheets[self.sheet_num - 1]]
        wb.active = sheetname
        nrows = sheetname.max_row
        ncols = sheetname.max_column
        flag = 0
        for i in range(2, nrows + 1):
            if sheetname.cell(i, 1).value == caseid:
                sheetname.cell(i, ncols).value = actual
                sheetname.cell(i, ncols - 3).value = 1
                flag = 1
                break
        if flag == 0:
            print("用例不存在！")
        wb.save(datafile_p)
        wb.close()

    def writetoken(self, token, idnum):
        wb = openpyxl.load_workbook(datafile_p)
        sheets = wb.sheetnames
        sheetname = wb[sheets[self.sheet_num - 1]]
        wb.active = sheetname
        ncols = sheetname.max_column
        header = eval(sheetname.cell(idnum, ncols-5).value)
        header['token'] = token
        sheetname.cell(idnum, ncols - 5).value = str(header)
        wb.save(datafile_p)
        wb.close()

    def initexec(self):
        wb = openpyxl.load_workbook(datafile_p)
        sheets = wb.sheetnames
        sheetname = wb[sheets[self.sheet_num - 1]]
        wb.active = sheetname
        nrows = sheetname.max_row
        ncols = sheetname.max_column
        for i in range(2, nrows + 1):
            if sheetname.cell(i, ncols-3).value != 0:
                sheetname.cell(i, ncols-3).value = 0
            sheetname.cell(i, ncols).value = ''
        wb.save(datafile_p)
        wb.close()


if __name__ == '__main__':
    sheet = ReadWrite(1)
    print(sheet.read())
