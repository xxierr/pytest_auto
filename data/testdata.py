# 读写.xlsx文件类
import openpyxl
import os

class RWxlsx:
    def __init__(self, file,num):
        self.file = file
        self.num = num

    def read(self):
        wb = openpyxl.load_workbook(self.file)
        sheets = wb.sheetnames
        table = wb[sheets[self.num-1]]
        wb.active = table
        nrows = table.max_row
        ncols = table.max_column
        row_values = []
        rows = []
        for row in range(2, nrows + 1):
            for col in range(1, ncols + 1):
                cell_value = table.cell(row, col).value
                row_values.append(cell_value)
            rows.append(row_values)
            row_values = []
        return rows
        # for row in range(2, nrows + 1):
        #     cell_value1 = table.cell(row, 1).value
        #     cell_value2 = table.cell(row, 2).value
        #     values = (cell_value1, cell_value2)
        #     row_values.append(values)
        # return row_values

    def write(self,username,password):
        wb = openpyxl.load_workbook(self.file)
        sheets = wb.sheetnames
        table = wb[sheets[self.num-1]]
        wb.active = table
        print(wb.active)
        nrows = table.max_row
        for rows in range(nrows + 1, nrows + 2):
            table.cell(rows, 1).value = username
            table.cell(rows, 2).value = password
        wb.save(self.file)
        wb.close()
